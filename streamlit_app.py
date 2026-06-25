import streamlit as st
from streamlit_image_coordinates import streamlit_image_coordinates
from PIL import Image, ImageDraw
import time
import io
import datetime
import requests
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ------------------------------------------------------------------
# CONFIGURATION: Paste your unique Power Automate HTTP URL link here
# ------------------------------------------------------------------
POWER_AUTOMATE_URL = "https://default9b2f9cbe865b4df8a5848494d8c1ef.f6.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/47ed5cfbda7d44a3a9ca56f439adaac0/triggers/manual/paths/invoke?api-version=1"

# Set up page layout
st.set_page_config(page_title="AGS Roof Leak Mapper", layout="wide")
st.title("🏭 AGS Roof Leak Mapping Tool")
st.info("💡 Choose plant and click on the left floor view to add a leak point. Use the dashboard below to manage labels, export temporary files, or submit directly to SharePoint.")

# 1. Plant Selection with your specific names
plant = st.selectbox("Select Plant:", ['Cambridge - 07', 'Oshawa - 04', 'Windsor - 02'])

# --- 🌤️ LIVE OPEN-METEO WEATHER ENGINE ---
@st.cache_data(ttl=3600)
def get_real_weather_data(plant_name, target_date):
    """
    Queries Open-Meteo's Archive API using the exact user-provided coordinates
    to extract actual measured Total Daily Precipitation.
    """
    coordinates = {
        'Cambridge - 07': {"lat": 43.403449, "lon": -80.322832},
        'Oshawa - 04': {"lat": 43.876437, "lon": -78.848991},
        'Windsor - 02': {"lat": 42.286758, "lon": -83.016596}
    }
    
    loc = coordinates.get(plant_name, coordinates['Cambridge - 07'])
    date_str = target_date.strftime("%Y-%m-%d")
    
    url = "https://archive-api.open-meteo.com/v1/archive"
    params = {
        "latitude": loc["lat"],
        "longitude": loc["lon"],
        "start_date": date_str,
        "end_date": date_str,
        "daily": ["precipitation_sum"],
        "timezone": "America/New_York"
    }
    
    try:
        response = requests.get(url, params=params, timeout=5)
        if response.status_code == 200:
            data = response.json()
            if "daily" in data:
                precip_sum = data["daily"]["precipitation_sum"][0]
                p_out = f"{round(precip_sum, 1)} mm" if precip_sum is not None else "0.0 mm"
                return p_out
    except Exception:
        pass
        
    return "N/A"


# Image pathways
if plant == "Cambridge - 07":
    left_path = "data/CambridgeCAD.png"
    right_path = "data/Cambridge.png"
elif plant == "Oshawa - 04":
    left_path = "data/OshawaCAD.png"
    right_path = "data/Oshawa.png"
else:
    left_path = "data/WindsorCAD.png"
    right_path = "data/Windsor.png"

# 2. Load and Resize Images to a stable layout width (600px)
try:
    left_img = Image.open(left_path).convert("RGB")
    right_img = Image.open(right_path).convert("RGB")
    
    DISPLAY_WIDTH = 600
    
    ratio_left = DISPLAY_WIDTH / left_img.width
    height_left = int(left_img.height * ratio_left)
    left_resized = left_img.resize((DISPLAY_WIDTH, height_left))
    
    ratio_right = DISPLAY_WIDTH / right_img.width
    height_right = int(right_img.height * ratio_right)
    right_resized = right_img.resize((DISPLAY_WIDTH, height_right))

except FileNotFoundError:
    st.error("⚠️ Could not find the image files. Please ensure files exist inside the 'data/' folder.")
    st.stop()

# Initialize session state tracking list for the active plant
plant_key = plant.replace(' ', '_').replace('-', '_')

if f"leak_points_{plant_key}" not in st.session_state:
    st.session_state[f"leak_points_{plant_key}"] = []

if f"last_click_{plant_key}" not in st.session_state:
    st.session_state[f"last_click_{plant_key}"] = None

# 3. Prepare Image Overlays
left_display = left_resized.copy()
right_display = right_resized.copy()

draw_left = ImageDraw.Draw(left_display)
draw_right = ImageDraw.Draw(right_display)

# Draw all existing saved points onto both images with high-contrast text backing boxes
for pt in st.session_state[f"leak_points_{plant_key}"]:
    x, y = pt['x'], pt['y']
    custom_name = pt['name']
    
    # Floor Map CAD Drawing Details (White background container box)
    draw_left.ellipse((x - 6, y - 6, x + 6, y + 6), fill="red")
    text_pos_left = (x + 10, y - 6)
    bbox_left = draw_left.textbbox(text_pos_left, custom_name)
    padded_bbox_left = (bbox_left[0] - 4, bbox_left[1] - 2, bbox_left[2] + 4, bbox_left[3] + 2)
    draw_left.rectangle(padded_bbox_left, fill="white", outline="red", width=1)
    draw_left.text(text_pos_left, custom_name, fill="red")
    
    # Roof Aerial View Drawing Details (Charcoal dark container background box)
    draw_right.ellipse((x - 12, y - 12, x + 12, y + 12), outline="cyan", width=3)
    draw_right.ellipse((x - 3, y - 3, x + 3, y + 3), fill="red")
    text_pos_right = (x + 16, y - 8)
    bbox_right = draw_right.textbbox(text_pos_right, custom_name)
    padded_bbox_right = (bbox_right[0] - 4, bbox_right[1] - 2, bbox_right[2] + 4, bbox_right[3] + 2)
    draw_right.rectangle(padded_bbox_right, fill="#1A1A1A", outline="cyan", width=1)
    draw_right.text(text_pos_right, custom_name, fill="cyan")

# 4. Display Side-by-Side Views
st.write("---")
col1, col2 = st.columns([1, 1])

with col1:
    st.markdown(f"### 🗺️ Floor Map View — **{plant}**")
    
    clicked_coords = streamlit_image_coordinates(
        left_display,
        key=f"image_click_{plant_key}" 
    )
    
    if clicked_coords is not None and clicked_coords != st.session_state[f"last_click_{plant_key}"]:
        st.session_state[f"last_click_{plant_key}"] = clicked_coords
        
        existing_points = st.session_state[f"leak_points_{plant_key}"]
        if existing_points:
            current_serial = max(pt['serial'] for pt in existing_points) + 1
        else:
            current_serial = 1
            
        unique_timestamp_id = str(time.time()).replace(".", "")
        
        st.session_state[f"leak_points_{plant_key}"].append({
            'id': unique_timestamp_id,
            'serial': current_serial,
            'x': clicked_coords['x'],
            'y': clicked_coords['y'],
            'name': f"Leak Point {current_serial}",
            'start_date': datetime.date.today()
        })
        st.rerun()

with col2:
    st.markdown("### 🦅 Corresponding Roof View")
    st.image(right_display, use_container_width=False)


# --- LOCAL EXCEL GENERATION GENERATOR ROUTINE ---
def export_to_excel_with_images(points, left_img_obj, right_img_obj, plant_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leak Mapping Report"
    ws.views.sheetView[0].showGridLines = True
    
    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    navy_bold = Font(name="Calibri", size=11, bold=True, color="1F497D")
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
    )
    
    ws.merge_cells("A1:G1")
    ws["A1"] = f"AGS Leak Mapping Report - {plant_name}"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    ws.row_dimensions[1].height = 30
    
    headers = [
        "Point ID", "Custom Label", "Date Noticed", 
        "Precipitation (Day Noticed)", "Precipitation (Day Before)",
        "Coordinate X", "Coordinate Y"
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[3].height = 24
    
    start_row = 4
    for idx, pt in enumerate(points):
        current_row = start_row + idx
        ws.row_dimensions[current_row].height = 20
        
        date_noticed = pt.get('start_date', datetime.date.today())
        date_before = date_noticed - datetime.timedelta(days=1)
        
        p_noticed = get_real_weather_data(plant_name, date_noticed)
        p_before = get_real_weather_data(plant_name, date_before)
        
        c1 = ws.cell(row=current_row, column=1, value=f"#{pt['serial']}")
        c2 = ws.cell(row=current_row, column=2, value=pt['name'])
        c3 = ws.cell(row=current_row, column=3, value=date_noticed.strftime("%Y-%m-%d"))
        c4 = ws.cell(row=current_row, column=4, value=p_noticed)
        c5 = ws.cell(row=current_row, column=5, value=p_before)
        c6 = ws.cell(row=current_row, column=6, value=int(pt['x']))
        c7 = ws.cell(row=current_row, column=7, value=int(pt['y']))
        
        for cell in [c1, c2, c3, c4, c5, c6, c7]:
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    image_heading_row = start_row + len(points) + 2
    ws.cell(row=image_heading_row, column=1, value="🗺️ Final Marked Floor Map View").font = navy_bold
    ws.cell(row=image_heading_row, column=5, value="🦅 Final Corresponding Roof View").font = navy_bold
    
    left_buffer, right_buffer = io.BytesIO(), io.BytesIO()
    left_img_obj.resize((450, int(left_img_obj.height * (450/left_img_obj.width)))).save(left_buffer, format="JPEG")
    right_img_obj.resize((450, int(right_img_obj.height * (450/right_img_obj.width)))).save(right_buffer, format="JPEG")
    
    left_buffer.seek(0)
    right_buffer.seek(0)
    
    ws.add_image(OpenpyxlImage(left_buffer), f"A{image_heading_row + 1}")
    ws.add_image(OpenpyxlImage(right_buffer), f"E{image_heading_row + 1}")
    
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream


# --- 📋 TRACKING DASHBOARD ---
st.write("---")
st.subheader("📋 Saved Leak Records Dashboard")

points_list = st.session_state[f"leak_points_{plant_key}"]

if not points_list:
    st.info(f"💡 No leaks mapped yet for {plant}. Click on the left Floor Map image to begin pinning locations.")
else:
    st.info("💡 **Click to rename:** Click directly inside any text box below to customize labels and select dates.")
    
    grid_header1, grid_header2, grid_header3, grid_header4, grid_header5, grid_header6 = st.columns([1.0, 2.2, 1.8, 2.3, 2.3, 1.4])
    with grid_header3:
        st.markdown("**📅 Date Noticed**")
    with grid_header4:
        st.markdown("**🌦️ Precipitation (Day Noticed)**")
    with grid_header5:
        st.markdown("**🌦️ Precipitation (Day Before)**")

    for index, point in enumerate(points_list):
        edit_col1, edit_col2, edit_col3, edit_col4, edit_col5, edit_col6 = st.columns([1.0, 2.2, 1.8, 2.3, 2.3, 1.4])
        
        with edit_col1:
            st.write(f"**#{point['serial']}**")
            
        with edit_col2:
            new_name = st.text_input(
                "Rename label:", 
                value=point['name'], 
                key=f"rename_{plant_key}_{point['id']}",
                label_visibility="collapsed"
            )
            if new_name != point['name']:
                st.session_state[f"leak_points_{plant_key}"][index]['name'] = new_name
                st.rerun()
                
        with edit_col3:
            chosen_date = st.date_input(
                "Date Leak Noticed",
                value=point.get('start_date', datetime.date.today()),
                max_value=datetime.date.today(),
                key=f"date_{plant_key}_{point['id']}",
                label_visibility="collapsed"
            )
            if chosen_date != point.get('start_date'):
                st.session_state[f"leak_points_{plant_key}"][index]['start_date'] = chosen_date
                st.rerun()
        
        day_before_date = chosen_date - datetime.timedelta(days=1)
        
        with edit_col4:
            p_noticed = get_real_weather_data(plant, chosen_date)
            st.markdown(f"**{p_noticed}**")
            
        with edit_col5:
            p_before = get_real_weather_data(plant, day_before_date)
            st.markdown(f"**{p_before}**")
                
        with edit_col6:
            if st.button("🗑️ Delete", key=f"del_{plant_key}_{point['id']}", use_container_width=True):
                st.session_state[f"leak_points_{plant_key}"].pop(index)
                st.rerun()

# --- 🚀 SUBMIT REPORT TO SHAREPOINT AND DOWNLOAD SECTION ---
if points_list:
    st.write("---")
    sub_col1, sub_col2 = st.columns([1, 1])
    
    with sub_col1:
        st.markdown("### 🚀 Production SharePoint Save")
        if st.button("Report Leaks & Save to SharePoint Master", use_container_width=True, type="primary"):
            if POWER_AUTOMATE_URL == "YOUR_POWER_AUTOMATE_HTTP_URL_HERE":
                st.error("⚠️ Please insert your actual generated Power Automate URL string at the top of the script code!")
            else:
                with st.spinner("Streaming leak records directly into SharePoint Master Table..."):
                    success_count = 0
                    
                    for pt in points_list:
                        d_noticed = pt.get('start_date', datetime.date.today())
                        d_before = d_noticed - datetime.timedelta(days=1)
                        
                        p_not = get_real_weather_data(plant, d_noticed)
                        p_bef = get_real_weather_data(plant, d_before)
                        
                        payload = {
                            "Plant": plant,
                            "Serial": int(pt['serial']),
                            "Label": pt['name'],
                            "DateNoticed": d_noticed.strftime("%Y-%m-%d"),
                            "PrecipNoticed": p_not,
                            "PrecipBefore": p_bef,
                            "CoordinateX": int(pt['x']),
                            "CoordinateY": int(pt['y']),
                            "Timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        }
                        
                        response = requests.post(POWER_AUTOMATE_URL, json=payload)
                        if response.status_code in [200, 202]:
                            success_count += 1
                            
                    if success_count == len(points_list):
                        st.success(f"🎉 Successfully saved {success_count} logs to your Master Excel Sheet!")
                        st.session_state[f"leak_points_{plant_key}"] = []
                        st.rerun()
                    else:
                        st.error("⚠️ Communication error occurred saving records to SharePoint.")

    with sub_col2:
        st.markdown("### 📥 Local Standalone Backup")
        excel_data = export_to_excel_with_images(points_list, left_display, right_display, plant)
        st.download_button(
            label="Download Data & Maps to Local Excel (.xlsx)",
            data=excel_data,
            file_name=f"AGS_Leak_Report_{plant_key}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
